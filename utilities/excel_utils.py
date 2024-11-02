"""
Module for handling Excel file operations.
This module provides utility functions to read and write data to Excel files
using the openpyxl library. It includes functions to get row and column counts,
read cell data, set cell data, and retrieve data as lists with optional filtering.
"""

import openpyxl


def get_row_count(path: str, sheet_name: str) -> int:
    """
    Returns the total number of rows in the specified sheet of the Excel file.

    Args:
        path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet from which to count rows.

    Returns:
        int: The total number of rows in the specified sheet.
    """
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(path: str, sheet_name: str) -> int:
    """
    Returns the total number of columns in the specified sheet of the Excel file.

    Args:
        path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet from which to count columns.

    Returns:
        int: The total number of columns in the specified sheet.
    """
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    return sheet.max_column


def get_cell_data(path: str, sheet_name: str, row_number: int, column_number: int):
    """
    Returns the data from a specific cell in the specified sheet of the Excel file.

    If the cell data is not a string, the returned value will be in a tuple or list.
    This is important for testing purposes, where each character is used as input data.

    Args:
        path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet from which to get cell data.
        row_number (int): The row number of the cell (1-indexed).
        column_number (int): The column number of the cell (1-indexed).

    Returns:
        The value of the cell, which may be a string, tuple, or list.
    """
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    cell_data = sheet.cell(row=row_number, column=column_number).value
    return cell_data


def set_cell_data(path: str, sheet_name: str, row_number: int, column_number: int, data):
    """
    Sets the data in a specific cell in the specified sheet of the Excel file.

    Args:
        path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet where the cell data should be set.
        row_number (int): The row number of the cell (1-indexed).
        column_number (int): The column number of the cell (1-indexed).
        data: The data to set in the specified cell.
    """
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row_number, column=column_number).value = data
    workbook.save(path)


def get_data_as_list(path: str, sheet_name: str, flag: bool = None) -> list:
    """
    Reads data from an Excel sheet and optionally filters it by a 'valid_flag' column.

    Args:
        path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet.
        flag (bool, optional): If True, returns only rows where 'valid_flag' is True.
                               If False, returns only rows where 'valid_flag' is False.
                               If None, returns all rows.

    Returns:
        list: A list of lists, where each inner list represents a row of data.
    """
    final_list = []
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    col_count = sheet.max_column

    for r in range(2, row_count + 1):  # Start from the second row to skip headers
        row_data = []
        for c in range(1, col_count + 1):
            row_data.append(sheet.cell(row=r, column=c).value)

        # Check the flag in the first column (index 0)
        if flag is None or row_data[0] == flag:
            # Append the data excluding the first column
            final_list.append(row_data[1:])

    return final_list
